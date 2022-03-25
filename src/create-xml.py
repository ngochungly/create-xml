import sys
# import xml.etree.ElementTree as ET
from openpyxl import load_workbook

# ===== Read .xlsx file =====
# filepath is the first argument
wb = load_workbook(sys.argv[1])
# Select sheet
ws = wb['Sheet1']

# Get rows and convert to a list
row_list = list(ws.iter_rows(min_row=2, max_row=4, min_col=3, max_col=None, values_only=False))

data = []

# Loop thru list
for row in row_list:
    row_str = ''
    row_dict = {}

    # Create a dict for each row
    for i in range(len(row)):
        key = row_list[0][i].value
        if (row[i].value != None):
            value = row[i].value
        else:
            value = ''
        #row_str += ' | {0}'.format(value)
        row_dict[key] = value

    # print row data
    # print(row_str)
    # Add row data to a list
    data.append(row_dict)

# Remove header row
data.pop(0)

# Create rubriques list
rubriques_list = []
for d in data:
    r = (d['Rubrique ID'], d['Rubrique Label'])
    if r not in rubriques_list:
        rubriques_list.append(r)

print('\n')
print(data)
print('\n')

# ===== Create .xml file =====
catalog_id = "155"
nom = "MCNC"
email = "replay030303@mcnc.tv"
xml_data_str = '<?xml version="1.0" encoding="utf-8"?>\n'

# Create root element
xml_data_str += '<catalog id="' + catalog_id + '">\n'

# Create child element
xml_data_str += '  <export>\n'
xml_data_str += '    <auteur>\n'
xml_data_str += '      <nom>' + nom + '</nom>\n'
xml_data_str += '      <email>' + email + '</email>\n'
xml_data_str += '    </auteur>\n'
xml_data_str += '  </export>\n\n'

# Add data from .xlsx file
xml_data_str += '  <programmes>\n'

for d in data:
    xml_data_str += '    <programme id="' + str(d['Program ID']) + '">\n'
    xml_data_str += '      <titre>' + str(d['Program Title']) + '</titre>\n'
    xml_data_str += '      <saison>' + str(d['Program Season']) + '</saison>\n'
    xml_data_str += '      <episode>' + str(d['Program Episode number']) + '</episode>\n'
    xml_data_str += '      <synopsis>' + str(d['Synopsis']) + '</synopsis>\n'
    xml_data_str += '      <annee>' + str(d['Production year']) + '</annee>\n'
    xml_data_str += '      <pays>' + str(d['Production country']) + '</pays>\n'
    xml_data_str += '      <duree>' + str(d['Duration (minutes)']) + '</duree>\n'

    xml_data_str += '      <premdiftv date="' + str(d['Première diffusion']) + '"/>\n'

    xml_data_str += '      <publications>\n'
    dp_str = str(d['Durée de publication'].days*24) + ':' + str(int(d['Durée de publication'].seconds/60)) + ':00'
    xml_data_str += '        <publication debut="' + str(d['Début de publication']) + '" duree="' + dp_str + '"/>\n'
    xml_data_str += '      </publications>\n'

    xml_data_str += '      <csa id="' + str(d['CSA ID']) + '"/>\n'

    xml_data_str += '      <rubriques>\n'
    xml_data_str += '        <rubrique id="' + str(d['Rubrique ID']) + '"/>\n'
    xml_data_str += '      </rubriques>\n'

    xml_data_str += '    </programme>\n\n'

xml_data_str += '  </programmes>\n\n'

xml_data_str += '  <correspondances>\n'
xml_data_str += '    <rubriques>\n'
for r in rubriques_list:
    xml_data_str += '      <rubrique id="' + str(r[0]) + '" label="' + str(r[1]) + '" rank="3"/>\n'
xml_data_str += '    </rubriques>\n'
xml_data_str += '  </correspondances>\n'


xml_data_str += '</catalog>\n'

print(xml_data_str)

# Write xml data to file
f = open('../data/output.xml', 'w', encoding='utf-8')
f.write(xml_data_str)
f.close()


