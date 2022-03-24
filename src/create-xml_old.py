from gettext import Catalog
from pydoc import synopsis
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl import load_workbook

# ===== Read .xlsx file =====
wb = load_workbook('../data/input.xlsx')
# Select sheet
ws = wb['Sheet1']

# Get rows and convert to a list
row_list = list(ws.iter_rows(min_row=2, max_row=4, min_col=3, max_col=None, values_only=False))

print('\n')
data = []

# Loop thru list
for row in row_list:
    row_str = ''
    row_dict = {}

    # Create a dict for each row
    for i in range(len(row)):
        key = row_list[0][i].value
        if (row[i].value != None):
            value = row[i].value
        else:
            value = ''
        row_str += ' | {0}'.format(value)
        row_dict[key] = value

    # print row data
    print(row_str)
    # Add row data to a list
    data.append(row_dict)

print('\n')
data.pop(0)
print(data)

# ===== Create .xml file =====
# Create root element
catalog = ET.Element('catalog')
catalog.set('id','xxx')

# Create child element
export = ET.SubElement(catalog, 'export')
auteur = ET.SubElement(export, 'auteur')
nom = ET.SubElement(auteur, 'nom')
nom.text = 'MCNC'
email = ET.SubElement(auteur, 'email')
email.text = 'replay030303@mcnc.tv'

# Add data from .xlsx file
programmes = ET.SubElement(catalog, 'programmes')
for d in data:
    programe = ET.SubElement(programmes, 'programme')
    programe.set('id', d['Program ID'])

    titre = ET.SubElement(programe, 'titre')
    titre.text = d['Program Title']

    saison = ET.SubElement(programe, 'saison')
    saison.text = d['Program Season']

    episode = ET.SubElement(programe, 'episode')
    episode.text = str(d['Program Episode number'])

    Synopsis = ET.SubElement(programe, 'synopsis')
    Synopsis.text = d['Synopsis']

    annee = ET.SubElement(programe, 'annee')
    annee.text = str(d['Production year'])

    pays = ET.SubElement(programe, 'pays')
    pays.text = d['Production country']

    duree = ET.SubElement(programe, 'duree')
    duree.text = str(d['Duration (minutes)'])

    premdiftv = ET.SubElement(programe, 'premdiftv')
    premdiftv.set('date', str(d['Première diffusion']))

    publications = ET.SubElement(programe, 'publications')
    publication = ET.SubElement(publications, 'publication')
    publication.set('debut', str(d['Début de publication']))
    publication.set('duree', str(d['Durée de publication']))

    csa = ET.SubElement(programe, 'csa')
    csa.set('id', str(d['CSA ID']))

    rubriques = ET.SubElement(programe, 'rubriques')
    rubrique = ET.SubElement(rubriques, 'rubrique')
    rubrique.set('id', d['Rubrique ID'])

correspondances = ET.SubElement(catalog, 'correspondances')
rubriques = ET.SubElement(correspondances, 'rubriques')
rubrique = ET.SubElement(rubriques, 'rubrique')
rubrique.set('id', 'ID3')
rubrique.set('label', 'Films et Séries')
rubrique.set('rank', '3')

# Write to file
tree = ET.ElementTree(element=catalog)
tree.write('../data/output.xml')



